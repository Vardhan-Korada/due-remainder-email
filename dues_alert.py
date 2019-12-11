import openpyxl
import smtplib
import sys
unpaid = {}
wb = openpyxl.load_workbook("/home/vardhan/Downloads/automate_online-materials/duesRecords.xlsx")
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row = 1, column = lastCol).value
for r in range(2,sheet.max_row+1):
    payment = sheet.cell(row = r,column=lastCol).value
    if payment != "paid":
        name = sheet.cell(row = r,column=1).value
        email = sheet.cell(row = r,column =2).value
        unpaid[name] = email
smtpObj = smtplib.SMTP("smtp.gamil.com",587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login("ncisanthonydinnozo@gmail.com",input("Enter your password: "))
for name,email in unpaid.items():
    body = "Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank You" %(latestMonth,name,latestMonth)
    print("Sending email to %s...." %email)
    sendmailStatus = smtpObj.sendmail("ncisanthonydinnozo@gmail.com",email,body)
    if sendmailStatus != {}:
        print("There was a problem sending email to %s : %s " %(email,sendmailStatus))
smtpObj.quit()

